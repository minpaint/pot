# directory/admin/quiz_admin.py
from django.contrib import admin
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.urls import reverse, path
from django.utils.safestring import mark_safe
from django.conf import settings
from django import forms
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from tablib import Dataset
from nested_admin import NestedModelAdmin, NestedTabularInline
from directory.models import (
    QuizCategory, QuizCategoryOrder, Quiz, Question, Answer, QuizAttempt, UserAnswer, QuizAccessToken
)
from directory.resources.quiz import QuizQuestionResource


class QuizAdminForm(forms.ModelForm):
    """Форма для админки квизов с правильной кодировкой"""
    class Meta:
        model = Quiz
        fields = '__all__'
        widgets = {
            'title': forms.TextInput(attrs={'style': 'width: 80%;'}),
            'description': forms.Textarea(attrs={'rows': 4, 'style': 'width: 80%;'}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Явно указываем кодировку для всех текстовых полей
        for field_name, field in self.fields.items():
            if isinstance(field.widget, (forms.TextInput, forms.Textarea)):
                field.widget.attrs.update({
                    'accept-charset': 'UTF-8',
                })


class QuizCategoryOrderInline(admin.TabularInline):
    """Inline для управления разделами в экзамене с порядком"""
    model = QuizCategoryOrder
    extra = 1
    fields = ['category', 'order']
    ordering = ['order']
    verbose_name = _("Раздел в экзамене")
    verbose_name_plural = _("Разделы экзамена (с порядком)")


class AnswerInline(NestedTabularInline):
    """Inline для ответов на вопрос"""
    model = Answer
    extra = 4
    min_num = 2
    max_num = 10
    fields = ['answer_text', 'is_correct', 'order']
    ordering = ['order']

    class Media:
        css = {
            'all': ('admin/css/forms.css',)
        }


@admin.register(QuizCategory)
class QuizCategoryAdmin(admin.ModelAdmin):
    """Админка для разделов экзаменов"""
    list_display = ['name', 'order', 'questions_count_display', 'images_count_display', 'is_active', 'created']
    list_filter = ['is_active', 'created']
    search_fields = ['name', 'description']
    list_editable = ['order', 'is_active']
    ordering = ['order', 'name']

    fieldsets = (
        (_('Основная информация'), {
            'fields': ('name', 'description')
        }),
        (_('Настройки'), {
            'fields': ('order', 'is_active')
        }),
    )

    def questions_count_display(self, obj):
        """Отображение количества вопросов"""
        count = obj.get_questions_count()
        return format_html('<strong>{}</strong> вопросов', count)
    questions_count_display.short_description = _('Всего вопросов')

    def images_count_display(self, obj):
        """Отображение количества вопросов с изображениями"""
        total = obj.get_questions_count()
        with_images = obj.questions.filter(is_active=True).exclude(image='').count()

        if with_images == 0:
            color = '#999'
            icon = 'fa-times-circle'
        elif with_images == total:
            color = '#4caf50'
            icon = 'fa-check-circle'
        else:
            color = '#ff9800'
            icon = 'fa-exclamation-circle'

        return format_html(
            '<span style="color: {};">'
            '<i class="fas {}"></i> <strong>{}</strong> из {}'
            '</span>',
            color, icon, with_images, total
        )
    images_count_display.short_description = _('С изображениями')

    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Добавляем кнопку импорта в контекст"""
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().change_view(request, object_id, form_url, extra_context=extra_context)


@admin.register(Quiz)
class QuizAdmin(admin.ModelAdmin):
    """Админка для экзаменов"""
    form = QuizAdminForm
    list_display = [
        'title', 'categories_display',
        'exam_total_questions', 'exam_time_limit', 'exam_allowed_incorrect',
        'assigned_users_count', 'attempts_count', 'is_active', 'created'
    ]
    list_filter = ['is_active', 'categories', 'created']
    search_fields = ['title', 'description']
    list_editable = ['is_active']
    ordering = ['-created']
    filter_horizontal = ['assigned_users']
    inlines = [QuizCategoryOrderInline]

    fieldsets = (
        (_('Основная информация'), {
            'fields': ('title', 'description')
        }),
        (_('Назначение пользователям'), {
            'fields': ('assigned_users',),
            'description': _('Выберите пользователей, которым доступен этот экзамен. Оставьте пустым для доступа всем.'),
        }),
        (_('Настройки прохождения'), {
            'fields': ('random_order', 'show_correct_answer', 'allow_skip')
        }),
        (_('Настройки экзамена'), {
            'fields': (
                'exam_total_questions',
                'exam_time_limit',
                'exam_allowed_incorrect',
                'use_adaptive_selection'
            ),
            'description': _(
                'Настройки экзамена: вопросы распределяются равномерно по всем разделам. '
                'Адаптивный подбор учитывает прогресс пользователя при выборе конкретных вопросов.'
            ),
        }),
        (_('Статистика'), {
            'fields': ('attempts_count', 'is_active'),
            'classes': ('collapse',)
        }),
    )

    def categories_display(self, obj):
        """Отображение разделов"""
        categories = obj.categories.all()
        count = categories.count()
        if count == 0:
            return format_html('<span style="color: gray;">Нет разделов</span>')
        names = ', '.join([cat.name for cat in categories[:3]])
        if count > 3:
            names += f'... (+{count - 3})'
        return format_html('<span title="{}">{}</span>', names, names)
    categories_display.short_description = _('Разделы')

    def assigned_users_count(self, obj):
        """Количество назначенных пользователей"""
        count = obj.assigned_users.count()
        if count == 0:
            return format_html('<span style="color: gray;">Доступен всем</span>')
        return format_html('<strong>{}</strong> пользователей', count)
    assigned_users_count.short_description = _('Назначено')


class HasImageFilter(admin.SimpleListFilter):
    """Фильтр для вопросов с изображениями"""
    title = _('наличие изображения')
    parameter_name = 'has_image'

    def lookups(self, request, model_admin):
        return (
            ('yes', _('С изображением')),
            ('no', _('Без изображения')),
        )

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.exclude(image='')
        if self.value() == 'no':
            return queryset.filter(image='')
        return queryset


@admin.register(Question)
class QuestionAdmin(NestedModelAdmin):
    """Админка для вопросов с использованием nested_admin"""
    list_display = ['question_number_display', 'question_text_short', 'category', 'image_preview', 'answers_count', 'is_active', 'created']
    list_filter = ['category', HasImageFilter, 'is_active', 'created']  # Добавлен фильтр по изображениям
    search_fields = ['question_text', 'order']  # Добавлен поиск по номеру
    list_editable = ['is_active']
    ordering = ['category', 'order', 'id']
    inlines = [AnswerInline]
    change_list_template = "admin/directory/question/change_list.html"

    fieldsets = (
        (_('Основная информация'), {
            'fields': ('order', 'category', 'question_text'),
            'description': _('Номер вопроса (order) соответствует позиции в исходном документе при импорте')
        }),
        (_('Дополнительно'), {
            'fields': ('image', 'explanation', 'is_active')
        }),
    )

    def get_urls(self):
        """🔗 Добавляем кастомный URL для экспорта"""
        urls = super().get_urls()
        custom_urls = [
            path('export/', self.admin_site.admin_view(self.export_view), name='directory_question_export'),
        ]
        return custom_urls + urls

    def export_view(self, request):
        """📤 Экспорт вопросов викторины в Excel с выделением правильных ответов"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Фильтрация по категории (если указана)
        category_id = request.GET.get('category_id')

        if category_id:
            queryset = Question.objects.filter(category_id=category_id)
        else:
            queryset = Question.objects.all()

        queryset = queryset.select_related('category').prefetch_related('answers').order_by('category__order', 'order', 'id')

        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Вопросы викторины"

        # Заголовки
        headers = [
            'Номер вопроса', 'Раздел', 'Текст вопроса',
            'Ответ 1', 'Ответ 2', 'Ответ 3', 'Ответ 4', 'Ответ 5',
            'Ответ 6', 'Ответ 7', 'Ответ 8', 'Ответ 9', 'Ответ 10',
            'Пояснение', 'Путь к изображению', 'Есть изображение', 'Активен'
        ]

        # Стиль для заголовка
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Стиль для правильного ответа
        correct_answer_font = Font(bold=True, color='006100', size=11)  # Темно-зеленый текст
        correct_answer_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Светло-зеленый фон

        # Заполняем данные
        row_num = 2
        for question in queryset:
            answers = list(question.answers.order_by('order'))

            # Номер вопроса
            ws.cell(row=row_num, column=1).value = question.order

            # Раздел
            ws.cell(row=row_num, column=2).value = question.category.name if question.category else ''

            # Текст вопроса
            ws.cell(row=row_num, column=3).value = question.question_text
            ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical='top')

            # Ответы (до 10 вариантов)
            for idx in range(10):
                col_num = 4 + idx  # Колонки 4-13 для ответов
                cell = ws.cell(row=row_num, column=col_num)

                if idx < len(answers):
                    answer = answers[idx]
                    cell.value = answer.answer_text
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                    # Выделяем правильный ответ
                    if answer.is_correct:
                        cell.font = correct_answer_font
                        cell.fill = correct_answer_fill
                else:
                    cell.value = ''

            # Пояснение
            ws.cell(row=row_num, column=14).value = question.explanation
            ws.cell(row=row_num, column=14).alignment = Alignment(wrap_text=True, vertical='top')

            # Путь к изображению
            ws.cell(row=row_num, column=15).value = question.image.name if question.image else ''

            # Есть изображение
            ws.cell(row=row_num, column=16).value = 'Да' if question.image else 'Нет'

            # Активен
            ws.cell(row=row_num, column=17).value = 'Да' if question.is_active else 'Нет'

            row_num += 1

        # Настраиваем ширину столбцов
        ws.column_dimensions['A'].width = 12  # Номер вопроса
        ws.column_dimensions['B'].width = 30  # Раздел
        ws.column_dimensions['C'].width = 50  # Текст вопроса

        # Ответы
        for i in range(10):
            col_letter = get_column_letter(4 + i)
            ws.column_dimensions[col_letter].width = 40

        ws.column_dimensions['N'].width = 50  # Пояснение
        ws.column_dimensions['O'].width = 30  # Путь к изображению
        ws.column_dimensions['P'].width = 15  # Есть изображение
        ws.column_dimensions['Q'].width = 12  # Активен

        # Закрепляем первую строку (заголовки)
        ws.freeze_panes = 'A2'

        # Формируем имя файла
        if category_id:
            try:
                from directory.models import QuizCategory
                category = QuizCategory.objects.get(id=category_id)
                filename = f"quiz_questions_{category.name}.xlsx"
            except:
                filename = "quiz_questions.xlsx"
        else:
            filename = "quiz_questions_all.xlsx"

        # Создаем HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Сохраняем книгу в response
        wb.save(response)

        return response

    def question_number_display(self, obj):
        """Отображение номера вопроса из импорта"""
        if obj.image:
            # С изображением - синий бейдж с иконкой
            return format_html(
                '<span style="display: inline-flex; align-items: center; gap: 4px; background: #2196F3; color: white; '
                'padding: 4px 10px; border-radius: 12px; font-weight: 600; font-size: 13px;">'
                '<i class="fas fa-image" style="font-size: 11px;"></i> №{}'
                '</span>',
                obj.order
            )
        else:
            # Без изображения - серый бейдж
            return format_html(
                '<span style="display: inline-block; background: #757575; color: white; '
                'padding: 4px 10px; border-radius: 12px; font-weight: 600; font-size: 13px;">'
                '№{}'
                '</span>',
                obj.order
            )
    question_number_display.short_description = _('№')
    question_number_display.admin_order_field = 'order'  # Позволяет сортировку по этой колонке

    def question_text_short(self, obj):
        """Сокращенный текст вопроса"""
        return obj.question_text[:80] + '...' if len(obj.question_text) > 80 else obj.question_text
    question_text_short.short_description = _('Вопрос')

    def image_preview(self, obj):
        """Превью изображения"""
        if obj.image:
            return format_html(
                '<img src="{}" style="max-width: 100px; max-height: 60px;" />',
                obj.image.url
            )
        return '-'
    image_preview.short_description = _('Изображение')

    def answers_count(self, obj):
        """Количество ответов"""
        return obj.answers.count()
    answers_count.short_description = _('Ответов')


class UserAnswerInline(admin.TabularInline):
    """Inline для ответов пользователя"""
    model = UserAnswer
    extra = 0
    can_delete = False
    readonly_fields = ['question', 'selected_answer', 'is_correct', 'is_skipped', 'answered_at']
    fields = ['question', 'selected_answer', 'is_correct', 'is_skipped', 'answered_at']

    def has_add_permission(self, request, obj=None):
        return False


@admin.register(QuizAttempt)
class QuizAttemptAdmin(admin.ModelAdmin):
    """Админка для попыток прохождения"""
    list_display = [
        'id', 'user', 'quiz', 'status', 'score_display', 'passed_display',
        'progress_display', 'failure_reason_display', 'started_at', 'completed_at'
    ]
    list_filter = ['status', 'passed', 'quiz', 'started_at']
    search_fields = ['user__username', 'user__first_name', 'user__last_name', 'quiz__title']
    readonly_fields = [
        'started_at', 'completed_at', 'time_limit_seconds',
        'allowed_incorrect_answers', 'max_questions',
        'failure_reason', 'incorrect_answers'
    ]
    ordering = ['-started_at']
    inlines = [UserAnswerInline]

    fieldsets = (
        (_('Основная информация'), {
            'fields': ('quiz', 'user', 'status')
        }),
        (_('Результаты'), {
            'fields': (
                'total_questions', 'correct_answers', 'skipped_questions',
                'incorrect_answers', 'score_percentage', 'passed'
            )
        }),
        (_('Настройки попытки'), {
            'fields': ('time_limit_seconds', 'allowed_incorrect_answers', 'max_questions', 'failure_reason'),
            'classes': ('collapse',)
        }),
        (_('Время'), {
            'fields': ('started_at', 'completed_at')
        }),
    )

    def score_display(self, obj):
        """Отображение результата с цветом"""
        color = 'green' if obj.passed else 'red'
        score_text = f'{obj.score_percentage:.1f}%'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, score_text
        )
    score_display.short_description = _('Результат')

    def passed_display(self, obj):
        """Отображение статуса прохождения"""
        if obj.status != QuizAttempt.STATUS_COMPLETED:
            return '-'
        if obj.passed:
            return format_html('<span style="color: green;">✓ Пройден</span>')
        return format_html('<span style="color: red;">✗ Не пройден</span>')
    passed_display.short_description = _('Статус')

    def progress_display(self, obj):
        """Отображение прогресса"""
        return format_html(
            '{} / {} <span style="color: gray;">(пропущено: {})</span>',
            obj.correct_answers,
            obj.total_questions,
            obj.skipped_questions
        )
    progress_display.short_description = _('Прогресс')

    def failure_reason_display(self, obj):
        """Отображение причины завершения"""
        return obj.get_failure_reason_display() if obj.failure_reason else _('Успешно завершено')
    failure_reason_display.short_description = _('Причина завершения')

    def has_add_permission(self, request):
        """Запретить создание попыток через админку"""
        return False


@admin.register(UserAnswer)
class UserAnswerAdmin(admin.ModelAdmin):
    """Админка для ответов пользователей"""
    list_display = ['id', 'attempt_link', 'user_display', 'question_short', 'is_correct', 'is_skipped', 'answered_at']
    list_filter = ['is_correct', 'is_skipped', 'answered_at']
    search_fields = ['attempt__user__username', 'question__question_text']
    readonly_fields = ['attempt', 'question', 'selected_answer', 'answered_at']
    ordering = ['-answered_at']

    def has_add_permission(self, request):
        """Запретить создание ответов через админку"""
        return False

    def attempt_link(self, obj):
        """Ссылка на попытку"""
        url = reverse('admin:directory_quizattempt_change', args=[obj.attempt.id])
        return format_html('<a href="{}">Попытка #{}</a>', url, obj.attempt.id)
    attempt_link.short_description = _('Попытка')

    def user_display(self, obj):
        """Отображение пользователя"""
        return obj.attempt.user.username
    user_display.short_description = _('Пользователь')

    def question_short(self, obj):
        """Сокращенный текст вопроса"""
        text = obj.question.question_text
        return text[:50] + '...' if len(text) > 50 else text
    question_short.short_description = _('Вопрос')


@admin.register(QuizAccessToken)
class QuizAccessTokenAdmin(admin.ModelAdmin):
    """Админка для токенов доступа к экзаменам"""
    list_display = [
        'user', 'quiz', 'status_display', 'valid_from', 'valid_until',
        'is_used', 'used_at', 'access_link_display', 'created_by', 'created'
    ]
    list_filter = ['is_active', 'is_used', 'created', 'valid_from', 'valid_until']
    search_fields = ['user__username', 'user__first_name', 'user__last_name', 'quiz__title', 'description']
    readonly_fields = ['token', 'is_used', 'used_at', 'created', 'modified', 'access_url_display']
    ordering = ['-created']

    fieldsets = (
        (_('Основная информация'), {
            'fields': ('quiz', 'user', 'description')
        }),
        (_('Период действия'), {
            'fields': ('valid_from', 'valid_until'),
            'description': _('Укажите период, в течение которого токен будет действителен')
        }),
        (_('Ссылка для доступа'), {
            'fields': ('token', 'access_url_display'),
            'description': _('Скопируйте и отправьте эту ссылку пользователю')
        }),
        (_('Состояние'), {
            'fields': ('is_active', 'is_used', 'used_at'),
        }),
        (_('Служебная информация'), {
            'fields': ('created_by', 'created', 'modified'),
            'classes': ('collapse',)
        }),
    )

    def save_model(self, request, obj, form, change):
        """Автоматически устанавливаем created_by при создании"""
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def status_display(self, obj):
        """Отображение статуса токена с цветом"""
        is_valid, message = obj.is_valid()
        if is_valid:
            return format_html('<span style="color: green; font-weight: bold;">✓ {}</span>', message)
        return format_html('<span style="color: red;">✗ {}</span>', message)
    status_display.short_description = _('Статус')

    def access_link_display(self, obj):
        """Кнопка копирования ссылки"""
        if obj.pk:
            domain = settings.EXAM_SUBDOMAIN
            protocol = settings.EXAM_PROTOCOL
            full_url = f"{protocol}://{domain}{obj.get_access_url()}"
            return format_html(
                '<div style="display: flex; align-items: center; gap: 10px;">'
                '<input type="text" value="{}" id="token-url-{}" readonly style="width: 400px; padding: 5px;" />'
                '<button type="button" onclick="copyToClipboard(\'token-url-{}\')" '
                'style="padding: 5px 10px; cursor: pointer;">Копировать</button>'
                '</div>'
                '<script>'
                'function copyToClipboard(elementId) {{'
                '    var input = document.getElementById(elementId);'
                '    input.select();'
                '    document.execCommand(\'copy\');'
                '    alert(\'Ссылка скопирована в буфер обмена!\');'
                '}}'
                '</script>',
                full_url, obj.pk, obj.pk
            )
        return '-'
    access_link_display.short_description = _('Ссылка доступа')

    def access_url_display(self, obj):
        """Отображение полной ссылки (readonly)"""
        if obj.pk:
            domain = settings.EXAM_SUBDOMAIN
            protocol = settings.EXAM_PROTOCOL
            full_url = f"{protocol}://{domain}{obj.get_access_url()}"
            return format_html(
                '<div style="background: #f0f0f0; padding: 10px; border-radius: 5px; font-family: monospace;">{}</div>',
                full_url
            )
        return _('Токен будет сгенерирован после сохранения')
    access_url_display.short_description = _('Полная ссылка')

    def has_delete_permission(self, request, obj=None):
        """Разрешить удаление только неиспользованных токенов"""
        if obj and obj.is_used:
            return False
        return super().has_delete_permission(request, obj)
