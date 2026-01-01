"""
🏭 Admin для структурного подразделения без MPTT.
Отображает древовидное представление подразделений.
Использует универсальный миксин TreeViewMixin.
"""
from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db import transaction
from django.core.exceptions import PermissionDenied
from io import BytesIO
import datetime
import openpyxl
import xlrd
from directory.admin.mixins.tree_view import TreeViewMixin
from directory.models import (
    StructuralSubdivision,
    SubdivisionEmail,
    Department,
    DepartmentEmail,
    Organization,
)
from deadline_control.models import EmailSettings
from directory.forms.subdivision import StructuralSubdivisionForm


class SubdivisionEmailInline(admin.TabularInline):
    """
    📧 Inline для управления email-адресами подразделения.

    Позволяет добавлять несколько email для одного подразделения
    с указанием роли получателя (Главный инженер, Мастер и т.д.).
    """
    model = SubdivisionEmail
    extra = 1
    fields = ['email', 'description', 'is_active', 'created_at']
    readonly_fields = ['created_at']

    verbose_name = "Email для уведомлений"
    verbose_name_plural = "Email-адреса для уведомлений"

    def get_queryset(self, request):
        """Оптимизируем запросы"""
        qs = super().get_queryset(request)
        return qs.select_related('subdivision')

    class Media:
        css = {
            'all': ('admin/css/forms.css',)
        }

@admin.register(StructuralSubdivision)
class StructuralSubdivisionAdmin(TreeViewMixin, admin.ModelAdmin):
    """
    🏭 Админ-класс для модели StructuralSubdivision.
    Отображает древовидное представление: Организация → Подразделение.
    """
    form = StructuralSubdivisionForm
    inlines = [SubdivisionEmailInline]

    # Используем шаблон, специфичный для структурных подразделений
    change_list_template = "admin/directory/subdivision/change_list_tree.html"

    # ⚙️ Настройки дерева: здесь ключевой параметр model_name определяет, что URL будет формироваться как
    # 'admin:directory_structuralsubdivision_change'
    tree_settings = {
        'icons': {
            'organization': '🏢',
            'subdivision': '🏭',
            'no_subdivision': '🏗️',
            # Для подразделения уровней "department" и "item" не используются, поэтому можно задать любую иконку:
            'department': '📂',
            'item': '🏭',
        },
        'fields': {
            'name_field': 'name',                # Имя подразделения
            'organization_field': 'organization',# FK на Organization
            'subdivision_field': None,             # Нет вложенных подразделений
            'department_field': None,              # Нет уровня "отдел"
        },
        'display_rules': {
            'hide_empty_branches': False,
            'hide_no_subdivision_no_department': False
        },
        # 🔑 Ключевой параметр: правильное имя модели для формирования URL
        'model_name': 'structuralsubdivision'
    }

    list_display = ['name', 'short_name', 'organization']
    list_filter = ['organization']
    search_fields = ['name', 'short_name']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if not request.user.is_superuser and hasattr(request.user, 'profile'):
            allowed_orgs = request.user.profile.organizations.all()
            qs = qs.filter(organization__in=allowed_orgs)
        return qs

    # ================================
    # Дополнительные URL (экспорт/импорт email)
    # ================================
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'email-manage/',
                self.admin_site.admin_view(self.email_manage_view),
                name='directory_structuralsubdivision_email_manage',
            ),
            path(
                'email-export/',
                self.admin_site.admin_view(self.email_export_view),
                name='directory_structuralsubdivision_email_export',
            ),
            path(
                'email-import/',
                self.admin_site.admin_view(self.email_import_view),
                name='directory_structuralsubdivision_email_import',
            ),
        ]
        return custom_urls + urls

    def _accessible_organizations(self, request):
        if request.user.is_superuser:
            return Organization.objects.all()
        if hasattr(request.user, 'profile'):
            return request.user.profile.organizations.all()
        return Organization.objects.none()

    # --------
    # Управление Email (главная страница)
    # --------
    def email_manage_view(self, request):
        """Единая страница для управления email: выбор организации и действия"""
        org_qs = self._accessible_organizations(request)
        if not org_qs.exists():
            raise PermissionDenied("Нет доступных организаций")

        context = {
            'title': "Управление Email адресами",
            'organizations': org_qs.order_by('short_name_ru'),
        }

        # Обработка формы
        if request.method == 'POST':
            action = request.POST.get('action')
            org_id = request.POST.get('organization')

            # Если выбрана конкретная организация
            if org_id and org_id != 'all':
                try:
                    org = org_qs.get(id=org_id)
                    # Сохраняем в сессии для передачи в export/import
                    request.session['email_manage_org_id'] = org.id
                except Organization.DoesNotExist:
                    messages.error(request, "Организация не найдена или недоступна")
                    return redirect(request.path)
            else:
                # Все организации
                request.session['email_manage_org_id'] = 'all'

            # Перенаправляем на соответствующее действие
            if action == 'export':
                return redirect('admin:directory_structuralsubdivision_email_export')
            elif action == 'import':
                return redirect('admin:directory_structuralsubdivision_email_import')
            else:
                messages.error(request, "Выберите действие")

        return render(request, "admin/directory/subdivision/email_manage.html", context)

    # --------
    # Export
    # --------
    def email_export_view(self, request):
        org_qs = self._accessible_organizations(request)
        if not org_qs.exists():
            raise PermissionDenied("Нет доступных организаций для экспорта")

        # Получаем выбранную организацию из сессии
        selected_org_id = request.session.get('email_manage_org_id', 'all')
        if selected_org_id != 'all':
            try:
                org_qs = org_qs.filter(id=selected_org_id)
                if not org_qs.exists():
                    messages.error(request, "Выбранная организация недоступна")
                    return redirect('admin:directory_structuralsubdivision_email_manage')
            except (ValueError, TypeError):
                pass

        # Очищаем сессию после использования
        request.session.pop('email_manage_org_id', None)

        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "emails"
        ws.append([
            "target_level",
            "org_short_name",
            "subdivision_name",
            "department_name",
            "email",
            "description",
            "is_active",
        ])

        # Организационный уровень
        email_settings = EmailSettings.objects.filter(
            organization__in=org_qs
        ).select_related('organization')
        for settings in email_settings:
            for email in settings.get_recipient_list():
                ws.append([
                    "organization",
                    settings.organization.short_name_ru,
                    "",
                    "",
                    email,
                    "",
                    True,
                ])

        # Подразделения
        subdivisions = StructuralSubdivision.objects.filter(
            organization__in=org_qs
        ).select_related('organization').prefetch_related('notification_emails')
        for subdivision in subdivisions:
            for email_obj in subdivision.notification_emails.all():
                ws.append([
                    "subdivision",
                    subdivision.organization.short_name_ru,
                    subdivision.name,
                    "",
                    email_obj.email,
                    email_obj.description,
                    email_obj.is_active,
                ])
            if not subdivision.notification_emails.exists():
                ws.append([
                    "subdivision",
                    subdivision.organization.short_name_ru,
                    subdivision.name,
                    "",
                    "",
                    "",
                    True,
                ])

        # Отделы
        departments = Department.objects.filter(
            organization__in=org_qs
        ).select_related('organization', 'subdivision').prefetch_related('notification_emails')
        for dept in departments:
            if dept.notification_emails.exists():
                for email_obj in dept.notification_emails.all():
                    ws.append([
                        "department",
                        dept.organization.short_name_ru,
                        dept.subdivision.name if dept.subdivision else "",
                        dept.name,
                        email_obj.email,
                        email_obj.description,
                        email_obj.is_active,
                    ])
            else:
                ws.append([
                    "department",
                    dept.organization.short_name_ru,
                    dept.subdivision.name if dept.subdivision else "",
                    dept.name,
                    "",
                    "",
                    True,
                ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"email_export_{timestamp}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # --------
    # Import
    # --------
    def email_import_view(self, request):
        # Получаем выбранную организацию из сессии
        selected_org_id = request.session.get('email_manage_org_id', 'all')
        selected_org_name = "все организации"

        if selected_org_id != 'all':
            try:
                org = self._accessible_organizations(request).get(id=selected_org_id)
                selected_org_name = org.short_name_ru
            except (Organization.DoesNotExist, ValueError, TypeError):
                selected_org_id = 'all'

        context = {
            'title': "Импорт e-mail для организаций/подразделений/отделов",
            'selected_org_id': selected_org_id,
            'selected_org_name': selected_org_name,
        }
        if request.method == 'POST':
            uploaded = request.FILES.get('import_file')
            apply_changes = bool(request.POST.get('apply'))
            if not uploaded:
                messages.error(request, "Загрузите XLS или XLSX файл")
                return redirect(request.path)

            try:
                parsed = self._parse_workbook(uploaded, request)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect(request.path)

            errors = parsed['errors']
            summary = parsed['summary']

            if errors:
                messages.error(request, f"Найдено ошибок: {len(errors)}")

            if apply_changes and not errors:
                self._apply_changes(parsed['payload'])
                # Очищаем сессию после успешного импорта
                request.session.pop('email_manage_org_id', None)
                messages.success(
                    request,
                    f"Импорт выполнен: организаций {summary['organizations']}, "
                    f"подразделений {summary['subdivisions']}, отделов {summary['departments']}"
                )
                return redirect('admin:directory_structuralsubdivision_changelist')

            context.update({
                'errors': errors,
                'summary': summary,
                'apply_changes': apply_changes,
            })
        return render(request, "admin/directory/subdivision/email_import.html", context)

    def _parse_workbook(self, uploaded, request):
        filename = uploaded.name.lower()
        rows = []
        if filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(uploaded, read_only=True)
            sheet = wb.active
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                rows.append((idx, row))
        elif filename.endswith('.xls'):
            book = xlrd.open_workbook(file_contents=uploaded.read())
            sheet = book.sheet_by_index(0)
            for idx in range(sheet.nrows):
                rows.append((idx + 1, sheet.row_values(idx)))
        else:
            raise ValueError("Поддерживаются только XLSX и XLS файлы")

        if not rows:
            raise ValueError("Файл пуст")

        header = [str(col).strip().lower() if col is not None else '' for col in rows[0][1]]
        expected = [
            "target_level",
            "org_short_name",
            "subdivision_name",
            "department_name",
            "email",
            "description",
            "is_active",
        ]
        if [h for h in header if h] != expected[:len([h for h in header if h])]:
            raise ValueError("Неверный заголовок файла. Ожидаются столбцы: " + ", ".join(expected))

        accessible_orgs = {
            org.short_name_ru.lower(): org
            for org in self._accessible_organizations(request)
        }

        errors = []
        payload = {
            'organizations': {},
            'subdivisions': {},
            'departments': {},
        }
        seen = {
            'organizations': {},
            'subdivisions': {},
            'departments': {},
        }

        def parse_bool(val):
            if val in (None, ''):
                return True
            if isinstance(val, bool):
                return val
            str_val = str(val).strip().lower()
            return str_val in ['1', 'true', 'yes', 'y', 'да', 'on', 'истина']

        for row_num, row_vals in rows[1:]:
            values = list(row_vals) + [''] * (7 - len(row_vals))
            target_level = (values[0] or '').strip().lower()
            org_short = (values[1] or '').strip()
            subdivision_name = (values[2] or '').strip()
            department_name = (values[3] or '').strip()
            raw_email = (values[4] or '').strip().lower()
            description = (values[5] or '').strip()
            is_active = parse_bool(values[6])

            if not target_level and not org_short and not raw_email:
                continue

            if target_level not in ('organization', 'subdivision', 'department'):
                errors.append((row_num, "target_level должен быть organization/subdivision/department"))
                continue

            org = accessible_orgs.get(org_short.lower())
            if not org:
                errors.append((row_num, f"Организация '{org_short}' недоступна или не найдена"))
                continue

            # email обязателен только если явно задан, пустой означает очистку
            email_value = raw_email
            if email_value:
                try:
                    SubdivisionEmail(email=email_value, subdivision=None).clean()
                except Exception as exc:
                    errors.append((row_num, f"Некорректный email '{email_value}': {exc}"))
                    continue

            key_org = org.id

            if target_level == 'organization':
                bucket = payload['organizations'].setdefault(key_org, [])
                seen_set = seen['organizations'].setdefault(key_org, set())
                if email_value and email_value in seen_set:
                    continue
                if email_value:
                    seen_set.add(email_value)
                    bucket.append({'email': email_value})
                else:
                    bucket.append({'email': ''})
                continue

            subdivision = None
            if target_level in ('subdivision', 'department'):
                if subdivision_name:
                    subdivision = StructuralSubdivision.objects.filter(
                        organization=org,
                        name__iexact=subdivision_name,
                    ).first()
                    if not subdivision:
                        errors.append((row_num, f"Подразделение '{subdivision_name}' не найдено в {org.short_name_ru}"))
                        continue
                elif target_level == 'subdivision':
                    errors.append((row_num, "Не задан subdivision_name"))
                    continue

            if target_level == 'subdivision':
                bucket = payload['subdivisions'].setdefault(subdivision.id, [])
                seen_set = seen['subdivisions'].setdefault(subdivision.id, set())
                if email_value and email_value in seen_set:
                    continue
                seen_set.add(email_value)
                bucket.append({
                    'email': email_value,
                    'description': description,
                    'is_active': is_active,
                })
                continue

            if not department_name:
                errors.append((row_num, "Не задан department_name"))
                continue

            department_qs = Department.objects.filter(
                organization=org,
                name__iexact=department_name,
            )
            if subdivision is None:
                department_qs = department_qs.filter(subdivision__isnull=True)
            else:
                department_qs = department_qs.filter(subdivision=subdivision)

            department = department_qs.first()
            if not department:
                errors.append((row_num, f"Отдел '{department_name}' не найден"))
                continue

            bucket = payload['departments'].setdefault(department.id, [])
            seen_set = seen['departments'].setdefault(department.id, set())
            if email_value and email_value in seen_set:
                continue
            seen_set.add(email_value)
            bucket.append({
                'email': email_value,
                'description': description,
                'is_active': is_active,
            })

        summary = {
            'organizations': len(payload['organizations']),
            'subdivisions': len(payload['subdivisions']),
            'departments': len(payload['departments']),
        }

        return {'errors': errors, 'summary': summary, 'payload': payload}

    def _apply_changes(self, payload):
        with transaction.atomic():
            # Организации: полностью заменяем recipient_emails
            for org_id, entries in payload['organizations'].items():
                org = Organization.objects.get(id=org_id)
                settings = EmailSettings.get_settings(org)
                cleaned = [e['email'] for e in entries if e.get('email')]
                settings.recipient_emails = "\n".join(cleaned)
                settings.save()

            # Подразделения
            for subdivision_id, entries in payload['subdivisions'].items():
                SubdivisionEmail.objects.filter(subdivision_id=subdivision_id).delete()
                bulk = []
                for entry in entries:
                    email = entry.get('email')
                    if not email:
                        continue
                    bulk.append(SubdivisionEmail(
                        subdivision_id=subdivision_id,
                        email=email,
                        description=entry.get('description', ''),
                        is_active=entry.get('is_active', True),
                    ))
                if bulk:
                    SubdivisionEmail.objects.bulk_create(bulk)

            # Отделы
            for department_id, entries in payload['departments'].items():
                DepartmentEmail.objects.filter(department_id=department_id).delete()
                bulk = []
                for entry in entries:
                    email = entry.get('email')
                    if not email:
                        continue
                    bulk.append(DepartmentEmail(
                        department_id=department_id,
                        email=email,
                        description=entry.get('description', ''),
                        is_active=entry.get('is_active', True),
                    ))
                if bulk:
                    DepartmentEmail.objects.bulk_create(bulk)
