"""
📊 Сервис для единого импорта/экспорта справочников

Поддерживает импорт и экспорт из одного Excel-файла с несколькими листами:
- Структура (Position)
- Сотрудники (Employee)
- Оборудование (Equipment)

Использует существующие Resource классы из django-import-export.
"""
from typing import Dict, List, Optional, Any
from tablib import Dataset
from django.db import transaction
from django.core.exceptions import ValidationError

from directory.resources.organization_structure import OrganizationStructureResource
from directory.resources.employee import EmployeeResource
from deadline_control.resources.equipment import EquipmentResource
from directory.models import Position, Employee, Organization
from deadline_control.models import Equipment


# Словарь сопоставления имён листов и ресурсов
RESOURCE_MAPPING = {
    'Структура': OrganizationStructureResource,
    'Сотрудники': EmployeeResource,
    'Оборудование': EquipmentResource,
}

# Критические листы (при ошибке в них - откат всей транзакции)
CRITICAL_SHEETS = {'Структура'}

# Порядок обработки листов (важно!)
PROCESSING_ORDER = ['Структура', 'Сотрудники', 'Оборудование']


def parse_workbook(file_obj) -> Dict[str, Dataset]:
    """
    Парсит Excel-файл и возвращает словарь {имя_листа: Dataset}

    Args:
        file_obj: Файловый объект (UploadedFile)

    Returns:
        Dict[str, Dataset]: Словарь с данными по листам

    Raises:
        ValidationError: Если файл некорректный или не содержит поддерживаемых листов
    """
    try:
        # Определяем формат файла
        file_name = getattr(file_obj, 'name', 'file.xlsx')
        file_format = file_name.split('.')[-1].lower()

        if file_format not in ['xlsx', 'xls']:
            raise ValidationError('Поддерживаются только файлы XLSX и XLS')

        # Читаем файл
        file_content = file_obj.read()
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)  # Возвращаем указатель в начало на случай повторного чтения

        # Используем openpyxl для чтения листов
        import openpyxl
        from io import BytesIO

        workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        sheet_names = workbook.sheetnames

        # Проверяем наличие хотя бы одного поддерживаемого листа
        found_sheets = [name for name in sheet_names if name in RESOURCE_MAPPING]
        if not found_sheets:
            raise ValidationError(
                f'Файл не содержит ни одного поддерживаемого листа. '
                f'Ожидаются листы: {", ".join(RESOURCE_MAPPING.keys())}'
            )

        # Парсим каждый найденный лист
        datasets = {}
        for sheet_name in found_sheets:
            sheet = workbook[sheet_name]

            # Читаем данные листа
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Пропускаем полностью пустые строки
                if any(cell is not None and str(cell).strip() for cell in row):
                    data.append(row)

            if not data:
                continue  # Пустой лист - пропускаем

            # Создаём Dataset из данных
            dataset = Dataset()

            # Первая строка - заголовки
            if data:
                headers = [str(h) if h is not None else '' for h in data[0]]
                dataset.headers = headers

                # Остальные строки - данные
                for row_data in data[1:]:
                    # Преобразуем None в пустые строки
                    clean_row = [str(cell) if cell is not None else '' for cell in row_data]
                    dataset.append(clean_row)

            datasets[sheet_name] = dataset

        workbook.close()
        return datasets

    except ValidationError:
        raise
    except Exception as e:
        raise ValidationError(f'Ошибка при чтении файла: {str(e)}')


def dry_run_import(datasets: Dict[str, Dataset], organization: Optional[Organization] = None) -> Dict[str, Any]:
    """
    Выполняет предпросмотр импорта без сохранения в БД

    Args:
        datasets: Словарь {имя_листа: Dataset}
        organization: Организация для ограничения импорта (опционально)
                     Если указана - в файле можно не указывать org_short_name_ru,
                     будет использована эта организация

    Returns:
        Dict с результатами по каждому листу:
        {
            'sheet_name': str,
            'status': 'ok'|'warn'|'error',
            'total_rows': int,
            'errors': List[{'row': int, 'message': str}],
            'result': ImportResult object
        }
    """
    results = []

    # Обрабатываем листы в правильном порядке
    sorted_sheets = [name for name in PROCESSING_ORDER if name in datasets]

    for sheet_name in sorted_sheets:
        dataset = datasets[sheet_name]
        resource_class = RESOURCE_MAPPING[sheet_name]

        sheet_result = {
            'sheet_name': sheet_name,
            'status': 'ok',
            'total_rows': len(dataset),
            'errors': [],
            'result': None,
        }

        try:
            # Нормализуем заголовки (поддержка упрощённых названий колонок)
            _normalize_dataset_headers(dataset, sheet_name)

            # Если указана организация - подставляем её в строки без org_short_name_ru
            if organization:
                _apply_organization_to_dataset(dataset, organization)

            # Создаём ресурс
            resource = resource_class()

            # Выполняем dry_run в транзакции (для безопасности)
            with transaction.atomic():
                result = resource.import_data(dataset, dry_run=True, raise_errors=False)
                sheet_result['result'] = result

                # Собираем ошибки
                if result.has_errors():
                    sheet_result['status'] = 'error'

                    for row_error in result.row_errors():
                        row_number = row_error[0]
                        errors_list = row_error[1]

                        for error in errors_list:
                            sheet_result['errors'].append({
                                'row': row_number,
                                'message': str(error.error)
                            })

                elif result.has_validation_errors():
                    sheet_result['status'] = 'warn'

                    # Собираем validation errors из invalid_rows
                    for invalid_row in result.invalid_rows:
                        row_number = invalid_row.number

                        # Собираем ошибки по полям
                        for field_name, error_messages in invalid_row.field_specific_errors.items():
                            for msg in error_messages:
                                sheet_result['errors'].append({
                                    'row': row_number,
                                    'message': f'[{field_name}] {msg}'
                                })

                        # Собираем общие ошибки (не привязанные к полям)
                        for msg in invalid_row.non_field_specific_errors:
                            sheet_result['errors'].append({
                                'row': row_number,
                                'message': str(msg)
                            })

                # Принудительный откат (даже если нет ошибок - это dry_run)
                transaction.set_rollback(True)

        except Exception as e:
            sheet_result['status'] = 'error'
            sheet_result['errors'].append({
                'row': 0,
                'message': f'Критическая ошибка обработки листа: {str(e)}'
            })

        results.append(sheet_result)

    return {
        'sheets': results,
        'has_errors': any(r['status'] == 'error' for r in results),
        'has_critical_errors': any(
            r['status'] == 'error' and r['sheet_name'] in CRITICAL_SHEETS
            for r in results
        ),
        'ready_for_import': not any(
            (r['status'] == 'error' or (r['status'] == 'warn' and r['sheet_name'] in CRITICAL_SHEETS))
            for r in results
        ),
    }


def _normalize_dataset_headers(dataset: Dataset, sheet_name: str):
    """
    Нормализует названия колонок датасета для совместимости с упрощёнными форматами

    Args:
        dataset: Dataset для обработки
        sheet_name: Название листа ('Структура', 'Сотрудники', 'Оборудование')
    """
    if not dataset.headers:
        return

    # Словарь алиасов: упрощённое_название → полное_название
    ALIASES = {
        'Структура': {
            'subdivision': 'subdivision_name',
            'department': 'department_name',
        },
        'Сотрудники': {
            'subdivision': 'subdivision_name',
            'department': 'department_name',
            'position': 'position_name',
        },
        'Оборудование': {
            'subdivision': 'subdivision_name',
            'department': 'department_name',
        },
    }

    if sheet_name not in ALIASES:
        return

    aliases = ALIASES[sheet_name]

    # Заменяем упрощённые названия на полные
    new_headers = []
    for header in dataset.headers:
        header_str = str(header).strip() if header else ''
        # Если есть алиас - используем полное название
        if header_str in aliases:
            new_headers.append(aliases[header_str])
        else:
            new_headers.append(header)

    dataset.headers = new_headers


def _apply_organization_to_dataset(dataset: Dataset, organization: Organization):
    """
    Подставляет org_short_name_ru в строки датасета, если оно пустое
    Если колонки нет - добавляет её в начало датасета

    Args:
        dataset: Dataset для обработки
        organization: Организация для подстановки
    """
    if not dataset.headers:
        return

    # Если колонки org_short_name_ru нет - добавляем её в начало
    if 'org_short_name_ru' not in dataset.headers:
        # Сначала добавляем значение организации в начало каждой строки
        for row_data in dataset._data:
            row_data.insert(0, organization.short_name_ru)

        # Затем добавляем заголовок в начало (после изменения данных!)
        dataset.headers = ['org_short_name_ru'] + list(dataset.headers)

        return

    # Если колонка уже есть - заполняем пустые значения
    org_index = dataset.headers.index('org_short_name_ru')

    # Модифицируем внутренние данные dataset._data, т.к. dataset.dict - это генератор
    for row_data in dataset._data:
        # row_data - это список значений
        if org_index < len(row_data):
            # Если поле пустое или содержит только пробелы - подставляем организацию
            if not row_data[org_index] or not str(row_data[org_index]).strip():
                row_data[org_index] = organization.short_name_ru


@transaction.atomic
def commit_import(datasets: Dict[str, Dataset], organization: Optional[Organization] = None) -> Dict[str, Any]:
    """
    Выполняет реальный импорт с сохранением в БД (атомарная транзакция)

    Args:
        datasets: Словарь {имя_листа: Dataset}
        organization: Организация для ограничения импорта (опционально)
                     Если указана - в файле можно не указывать org_short_name_ru,
                     будет использована эта организация

    Returns:
        Dict с результатами по каждому листу:
        {
            'sheet_name': str,
            'status': 'success'|'error',
            'created': int,
            'updated': int,
            'errors': int,
            'skipped': int,
        }
    """
    results = []

    # Обрабатываем листы в правильном порядке
    sorted_sheets = [name for name in PROCESSING_ORDER if name in datasets]

    try:
        for sheet_name in sorted_sheets:
            dataset = datasets[sheet_name]
            resource_class = RESOURCE_MAPPING[sheet_name]

            sheet_result = {
                'sheet_name': sheet_name,
                'status': 'success',
                'created': 0,
                'updated': 0,
                'errors': 0,
                'skipped': 0,
            }

            try:
                # Нормализуем заголовки (поддержка упрощённых названий колонок)
                _normalize_dataset_headers(dataset, sheet_name)

                # Если указана организация - подставляем её в строки без org_short_name_ru
                if organization:
                    _apply_organization_to_dataset(dataset, organization)

                # Создаём ресурс
                resource = resource_class()

                # Выполняем импорт
                result = resource.import_data(dataset, dry_run=False, raise_errors=False)

                # Собираем статистику
                sheet_result['created'] = result.totals.get('new', 0)
                sheet_result['updated'] = result.totals.get('update', 0)
                sheet_result['errors'] = result.totals.get('error', 0)
                sheet_result['skipped'] = result.totals.get('skip', 0) + result.totals.get('invalid', 0)

                # Проверяем на ошибки
                if result.has_errors():
                    sheet_result['status'] = 'error'

                    # Если это критический лист - прерываем транзакцию
                    if sheet_name in CRITICAL_SHEETS:
                        raise ValidationError(
                            f'Импорт остановлен. Обнаружены ошибки в критическом листе "{sheet_name}". '
                            f'Ошибок: {sheet_result["errors"]}'
                        )

                # Validation errors не блокируют импорт, но логируются
                elif result.has_validation_errors():
                    sheet_result['status'] = 'warning'
                    # Невалидные строки пропускаются, но импорт продолжается

            except ValidationError:
                raise  # Пробрасываем ValidationError дальше для отката транзакции
            except Exception as e:
                sheet_result['status'] = 'error'
                sheet_result['errors'] = 1

                # Если это критический лист - прерываем
                if sheet_name in CRITICAL_SHEETS:
                    raise ValidationError(
                        f'Импорт остановлен. Критическая ошибка в листе "{sheet_name}": {str(e)}'
                    )

            results.append(sheet_result)

        return {
            'sheets': results,
            'success': all(r['status'] == 'success' for r in results),
            'total_created': sum(r['created'] for r in results),
            'total_updated': sum(r['updated'] for r in results),
            'total_errors': sum(r['errors'] for r in results),
        }

    except ValidationError as e:
        # Откатываем транзакцию и возвращаем ошибку
        transaction.set_rollback(True)
        return {
            'sheets': results,
            'success': False,
            'error_message': str(e),
        }


def export_all_to_workbook(organization: Optional[Organization] = None) -> bytes:
    """
    Экспортирует все три справочника в один Excel-файл с несколькими листами

    Args:
        organization: Организация для фильтрации данных (опционально)

    Returns:
        bytes: Содержимое Excel-файла
    """
    import openpyxl
    from io import BytesIO

    # Создаём новый workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    # Экспортируем каждый справочник
    for sheet_name in PROCESSING_ORDER:
        resource_class = RESOURCE_MAPPING[sheet_name]
        resource = resource_class()

        # Получаем queryset для экспорта
        if sheet_name == 'Структура':
            queryset = Position.objects.all()
        elif sheet_name == 'Сотрудники':
            queryset = Employee.objects.all()
        elif sheet_name == 'Оборудование':
            queryset = Equipment.objects.all()
        else:
            continue

        # Фильтруем по организации (если указана)
        if organization:
            queryset = queryset.filter(organization=organization)

        # Оптимизируем запрос
        if sheet_name == 'Структура':
            queryset = queryset.select_related('organization', 'subdivision', 'department')
        elif sheet_name == 'Сотрудники':
            queryset = queryset.select_related('organization', 'subdivision', 'department', 'position')
        elif sheet_name == 'Оборудование':
            queryset = queryset.select_related('organization', 'subdivision', 'department')

        # Экспортируем dataset
        dataset = resource.export(queryset)

        # Создаём лист в workbook
        ws = wb.create_sheet(title=sheet_name)

        # Записываем заголовки
        if dataset.headers:
            ws.append(dataset.headers)

        # Записываем данные
        for row in dataset:
            ws.append(row)

        # Автоматическая ширина колонок (для удобства)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
