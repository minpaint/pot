"""
📊 Сервис для импорта реестра сотрудников

Обрабатывает иерархический Excel-файл с сотрудниками:
- Организация берётся из файла или задаётся вручную
- Subdivision/Department парсятся из иерархических путей
- Position создаются с привязкой к subdivision/department
- Employee создаются со всеми связями
"""
from typing import Dict, List, Optional, Any, Tuple
from django.db import transaction
from django.core.exceptions import ValidationError
from datetime import datetime, date
import openpyxl

from directory.models import (
    Organization, StructuralSubdivision, Department,
    Position, Employee
)


class RegistryParseResult:
    """Результат парсинга файла реестра"""

    def __init__(self):
        self.organization: Optional[str] = None
        self.header_row: int = 0
        self.rows_data: List[Dict[str, Any]] = []
        self.errors: List[Dict[str, str]] = []

        # Статистика
        self.total_rows: int = 0
        self.employees_count: int = 0
        self.subdivisions_count: int = 0
        self.departments_count: int = 0
        self.positions_count: int = 0


class RegistryImportResult:
    """Результат импорта реестра"""

    def __init__(self):
        self.success: bool = False
        self.employees_created: int = 0
        self.employees_updated: int = 0
        self.subdivisions_created: int = 0
        self.departments_created: int = 0
        self.positions_created: int = 0
        self.errors: List[Dict[str, Any]] = []
        self.error_message: Optional[str] = None


def _clean_part(value: str) -> str:
    return " ".join(str(value).strip().split())


def _normalize_key(value: str) -> str:
    return _clean_part(value).lower()


def parse_subdivision_path(
    path: str,
    existing_subdivisions: Optional[Dict[str, str]] = None
) -> Tuple[str, Optional[str]]:
    """
    Парсит путь подразделения

    Правила:
    - Одинарная: "Управление" → (Subdivision="Управление", Department=None)
    - Двойная: "Филиал / Отдел" → (Subdivision="Филиал", Department="Отдел")

    Args:
        path: Путь подразделения из файла

    Returns:
        Tuple[subdivision_name, department_name]
    """
    parts = [_clean_part(p) for p in path.split('/')]
    parts = [p for p in parts if p]

    if len(parts) == 1:
        return parts[0], None
    if existing_subdivisions:
        first_key = _normalize_key(parts[0])
        if first_key in existing_subdivisions:
            subdivision_name = existing_subdivisions[first_key]
            department_name = ' / '.join(parts[1:]) if len(parts) > 1 else None
            return subdivision_name, department_name

        # Нет совпадения - отбрасываем первый элемент и берём остаток
        parts = parts[1:] if len(parts) > 1 else parts
        if len(parts) == 1:
            return parts[0], None
        return parts[0], ' / '.join(parts[1:])

    if len(parts) == 2:
        return parts[0], parts[1]
    # Если больше 2 уровней - всё кроме последнего в subdivision
    return parts[0], ' / '.join(parts[1:])


def parse_date(value) -> Optional[date]:
    """Парсит дату из различных форматов"""
    if not value:
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, str):
        # Пробуем разные форматы
        for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except (ValueError, TypeError):
                continue

    return None


def find_organization_in_file(ws) -> Optional[str]:
    """
    Ищет организацию в первых 20 строках файла

    Args:
        ws: Worksheet объект openpyxl

    Returns:
        Название организации или None
    """
    for row_idx in range(1, 21):
        for col_idx in range(1, 10):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and 'БЕЛВИЛЛЕСДЕН' in str(cell_value):
                # Очищаем от префикса "Организация:"
                org_str = str(cell_value).strip()
                if ':' in org_str:
                    return org_str.split(':', 1)[1].strip()
                else:
                    return org_str
    return None


def find_header_row(ws) -> Optional[int]:
    """
    Ищет строку с заголовками (содержащую "ФИО")

    Args:
        ws: Worksheet объект openpyxl

    Returns:
        Номер строки или None
    """
    for row_idx in range(1, 20):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 9)]
        if any('ФИО' in str(v).upper() if v else False for v in row_values):
            return row_idx
    return None


def parse_registry_file(file_obj, organization_override: Optional[Organization] = None) -> RegistryParseResult:
    """
    Парсит Excel-файл с реестром сотрудников

    Args:
        file_obj: Файловый объект (UploadedFile)
        organization_override: Организация для переопределения (если не указана - берём из файла)

    Returns:
        RegistryParseResult с распарсенными данными
    """
    result = RegistryParseResult()

    try:
        # Открываем файл
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active

        # Определяем организацию
        if organization_override:
            result.organization = organization_override.short_name_ru
        else:
            org_from_file = find_organization_in_file(ws)
            if not org_from_file:
                raise ValidationError(
                    'Организация не найдена в файле. '
                    'Укажите организацию вручную при импорте.'
                )
            result.organization = org_from_file

        # Ищем заголовки
        result.header_row = find_header_row(ws)
        if not result.header_row:
            raise ValidationError('Не найдена строка с заголовками (должна содержать "ФИО")')

        # Подготавливаем справочник существующих подразделений
        existing_subdivisions = None
        organization_for_lookup = organization_override
        if not organization_for_lookup and result.organization:
            organization_for_lookup = Organization.objects.filter(
                short_name_ru=result.organization
            ).first()

        if organization_for_lookup:
            existing_subdivisions = {}
            for name in StructuralSubdivision.objects.filter(
                organization=organization_for_lookup
            ).values_list('name', flat=True):
                key = _normalize_key(name)
                if key and key not in existing_subdivisions:
                    existing_subdivisions[key] = name

        # Парсим строки
        COL_SUBDIVISION = 3
        COL_POSITION = 4
        COL_FIO = 5
        COL_HIRE_DATE = 6
        COL_BIRTH_DATE = 7

        current_subdivision = None
        current_department = None

        for row_idx in range(result.header_row + 1, ws.max_row + 1):
            subdivision_raw = ws.cell(row=row_idx, column=COL_SUBDIVISION).value
            position = ws.cell(row=row_idx, column=COL_POSITION).value
            fio = ws.cell(row=row_idx, column=COL_FIO).value
            hire_date = ws.cell(row=row_idx, column=COL_HIRE_DATE).value
            birth_date = ws.cell(row=row_idx, column=COL_BIRTH_DATE).value

            result.total_rows += 1

            # Обновляем контекст подразделения
            if subdivision_raw:
                subdivision_path = str(subdivision_raw).strip()
                subdivision_name, department_name = parse_subdivision_path(
                    subdivision_path,
                    existing_subdivisions=existing_subdivisions
                )

                current_subdivision = subdivision_name
                current_department = department_name

            # Обрабатываем сотрудника
            if position and fio:
                # Валидация
                if not current_subdivision:
                    result.errors.append({
                        'row': row_idx,
                        'message': 'Сотрудник без подразделения',
                        'fio': str(fio)
                    })
                    continue

                row_data = {
                    'row_number': row_idx,
                    'subdivision': current_subdivision,
                    'department': current_department,
                    'position': str(position).strip(),
                    'fio': str(fio).strip(),
                    'hire_date': parse_date(hire_date),
                    'birth_date': parse_date(birth_date),
                }

                result.rows_data.append(row_data)
                result.employees_count += 1

        wb.close()

        # Подсчитываем уникальные элементы
        result.subdivisions_count = len(set(r['subdivision'] for r in result.rows_data))
        departments = set((r['subdivision'], r['department']) for r in result.rows_data if r['department'])
        result.departments_count = len(departments)

        positions = set((r['subdivision'], r['department'], r['position']) for r in result.rows_data)
        result.positions_count = len(positions)

    except ValidationError:
        raise
    except Exception as e:
        raise ValidationError(f'Ошибка при чтении файла: {str(e)}')

    return result


@transaction.atomic
def import_registry_data(
    parse_result: RegistryParseResult,
    organization: Organization,
    update_existing: bool = False
) -> RegistryImportResult:
    """
    Импортирует данные реестра в базу данных

    Args:
        parse_result: Результат парсинга файла
        organization: Организация для импорта
        update_existing: Обновлять существующих сотрудников

    Returns:
        RegistryImportResult с результатами импорта
    """
    result = RegistryImportResult()

    try:
        # Кеши для созданных объектов
        subdivisions_cache = {}  # subdivision_name -> StructuralSubdivision
        departments_cache = {}   # (subdivision_name, department_name) -> Department
        positions_cache = {}     # (subdivision_name, department_name, position_name) -> Position

        for row_data in parse_result.rows_data:
            try:
                # 1. Создаём/находим Subdivision
                subdivision_name = row_data['subdivision']
                if subdivision_name not in subdivisions_cache:
                    subdivision, created = StructuralSubdivision.objects.get_or_create(
                        name=subdivision_name,
                        organization=organization,
                        defaults={'short_name': subdivision_name}
                    )
                    subdivisions_cache[subdivision_name] = subdivision
                    if created:
                        result.subdivisions_created += 1
                else:
                    subdivision = subdivisions_cache[subdivision_name]

                # 2. Создаём/находим Department (если есть)
                department = None
                department_name = row_data['department']
                if department_name:
                    dept_key = (subdivision_name, department_name)
                    if dept_key not in departments_cache:
                        department, created = Department.objects.get_or_create(
                            name=department_name,
                            organization=organization,
                            subdivision=subdivision,
                            defaults={'short_name': department_name}
                        )
                        departments_cache[dept_key] = department
                        if created:
                            result.departments_created += 1
                    else:
                        department = departments_cache[dept_key]

                # 3. Создаём/находим Position (с привязкой к subdivision/department)
                position_name = row_data['position']
                pos_key = (subdivision_name, department_name, position_name)
                if pos_key not in positions_cache:
                    position, created = Position.objects.get_or_create(
                        position_name=position_name,
                        organization=organization,
                        subdivision=subdivision,
                        department=department,
                        defaults={
                            'internship_period_days': 0,
                            'is_responsible_for_safety': False,
                            'can_be_internship_leader': False,
                            'can_sign_orders': False,
                        }
                    )
                    positions_cache[pos_key] = position
                    if created:
                        result.positions_created += 1
                else:
                    position = positions_cache[pos_key]

                # 4. Создаём/обновляем Employee
                fio = row_data['fio']

                # Ищем существующего сотрудника
                existing_employee = Employee.objects.filter(
                    full_name_nominative=fio,
                    organization=organization
                ).first()

                if existing_employee and not update_existing:
                    # Пропускаем, если не обновляем
                    continue

                employee_data = {
                    'organization': organization,
                    'subdivision': subdivision,
                    'department': department,
                    'position': position,
                    'contract_type': 'standard',
                    'status': 'active',
                }

                # Добавляем даты если есть
                if row_data['hire_date']:
                    employee_data['hire_date'] = row_data['hire_date']
                    employee_data['start_date'] = row_data['hire_date']

                if row_data['birth_date']:
                    employee_data['date_of_birth'] = row_data['birth_date']

                if existing_employee:
                    # Обновляем существующего
                    for key, value in employee_data.items():
                        setattr(existing_employee, key, value)
                    existing_employee.save()
                    result.employees_updated += 1
                else:
                    # Создаём нового
                    Employee.objects.create(
                        full_name_nominative=fio,
                        **employee_data
                    )
                    result.employees_created += 1

            except Exception as e:
                result.errors.append({
                    'row': row_data['row_number'],
                    'fio': row_data['fio'],
                    'error': str(e)
                })

        result.success = len(result.errors) == 0

    except Exception as e:
        result.success = False
        result.error_message = f'Критическая ошибка при импорте: {str(e)}'
        raise

    return result


def dry_run_import(
    parse_result: RegistryParseResult,
    organization: Organization
) -> Dict[str, Any]:
    """
    Выполняет пробный импорт без сохранения в БД

    Args:
        parse_result: Результат парсинга
        organization: Организация для проверки

    Returns:
        Dict с результатами проверки
    """
    preview = {
        'organization': organization.short_name_ru,
        'total_employees': parse_result.employees_count,
        'total_rows': parse_result.total_rows,
        'subdivisions_count': parse_result.subdivisions_count,
        'departments_count': parse_result.departments_count,
        'positions_count': parse_result.positions_count,
        'errors': parse_result.errors,
        'sample_data': parse_result.rows_data[:20],  # Первые 20 для предпросмотра
    }

    return preview
