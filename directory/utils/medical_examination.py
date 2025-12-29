import os
import csv
import json
import datetime
import openpyxl
from io import BytesIO
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q

from deadline_control.models.medical_examination import MedicalExaminationType, HarmfulFactor, MedicalSettings
from deadline_control.models.medical_norm import (
    MedicalExaminationNorm,
    PositionMedicalFactor,
    EmployeeMedicalExamination
)
from directory.models.position import Position
from directory.models.employee import Employee


def update_medical_examination_statuses():
    """
    Обновляет статусы медосмотров сотрудников

    Автоматически меняет статусы медосмотров в зависимости от сроков:
    1. Для медосмотров, до окончания которых осталось меньше заданного периода - "Нужно выдать направление"
    2. Для просроченных медосмотров - "Просрочен"

    Обрабатывает каждую организацию отдельно с учетом её настроек.

    Returns:
        dict: Информация о количестве обновленных записей
    """
    from directory.models import Organization
    today = timezone.now().date()

    to_issue_count = 0
    expired_count = 0

    # Обрабатываем каждую организацию отдельно
    for organization in Organization.objects.all():
        settings = MedicalSettings.get_settings(organization)
        if not settings:
            continue

        # Определяем даты для изменения статусов
        issue_date = today + datetime.timedelta(days=settings.days_before_issue)

        # Находим все медосмотры со статусом "Пройден", у которых срок подходит к концу
        to_issue_exams = EmployeeMedicalExamination.objects.filter(
            employee__organization=organization,
            next_date__lte=issue_date,
            next_date__gt=today,
            status='completed'
        )

        # Меняем статус на "Нужно выдать направление"
        to_issue_count += to_issue_exams.update(status='to_issue')

        # Находим все просроченные медосмотры
        expired_exams = EmployeeMedicalExamination.objects.filter(
            employee__organization=organization,
            next_date__lt=today,
            status__in=['completed', 'to_issue']
        )

        # Меняем статус на "Просрочен"
        expired_count += expired_exams.update(status='expired')

    return {
        'to_issue_updated': to_issue_count,
        'expired_updated': expired_count,
        'timestamp': timezone.now()
    }


def import_medical_norms_from_file(file, skip_first_row=True, update_existing=False):
    """
    Импорт норм медосмотров из файла

    Args:
        file: Загруженный файл (xlsx или csv)
        skip_first_row: Пропустить первую строку (заголовки)
        update_existing: Обновлять существующие записи

    Returns:
        dict: Результаты импорта
    """
    filename = file.name.lower()
    total_rows = 0
    imported = 0
    updated = 0
    errors = []

    try:
        if filename.endswith('.xlsx'):
            data = import_from_xlsx(file, skip_first_row)
        elif filename.endswith('.csv'):
            data = import_from_csv(file, skip_first_row)
        else:
            return {'success': False, 'message': 'Неподдерживаемый формат файла'}

        with transaction.atomic():
            for row_idx, row in enumerate(data):
                total_rows += 1

                try:
                    # Ожидаемые столбцы: название_должности, код_фактора, переопределение_периодичности, примечания
                    if len(row) < 2:
                        errors.append(f"Строка {row_idx + 1 + (1 if skip_first_row else 0)}: Недостаточно данных")
                        continue

                    position_name = row[0].strip()
                    factor_code = row[1].strip()

                    periodicity_override = None
                    if len(row) > 2 and row[2]:
                        try:
                            periodicity_override = int(row[2])
                            if periodicity_override <= 0:
                                raise ValueError("Периодичность должна быть положительным числом")
                        except ValueError:
                            errors.append(f"Строка {row_idx + 1 + (1 if skip_first_row else 0)}: "
                                          f"Некорректное значение периодичности: {row[2]}")
                            continue

                    notes = row[3] if len(row) > 3 else ""

                    # Проверяем существование вредного фактора
                    try:
                        harmful_factor = HarmfulFactor.objects.get(short_name=factor_code)
                    except HarmfulFactor.DoesNotExist:
                        errors.append(f"Строка {row_idx + 1 + (1 if skip_first_row else 0)}: "
                                      f"Вредный фактор с кодом '{factor_code}' не найден")
                        continue

                    # Проверяем существование нормы
                    norm, created = MedicalExaminationNorm.objects.get_or_create(
                        position_name=position_name,
                        harmful_factor=harmful_factor,
                        defaults={
                            'periodicity_override': periodicity_override,
                            'notes': notes
                        }
                    )

                    if created:
                        imported += 1
                    elif update_existing:
                        norm.periodicity_override = periodicity_override
                        norm.notes = notes
                        norm.save()
                        updated += 1

                except Exception as e:
                    errors.append(f"Строка {row_idx + 1 + (1 if skip_first_row else 0)}: {str(e)}")

        return {
            'success': True,
            'total_rows': total_rows,
            'imported': imported,
            'updated': updated,
            'errors': errors,
        }

    except Exception as e:
        return {'success': False, 'message': f'Ошибка импорта: {str(e)}'}


def import_from_xlsx(file, skip_first_row):
    """Импорт данных из Excel-файла"""
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    data = []

    start_row = 1 if not skip_first_row else 2
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if any(cell is not None for cell in row):  # Пропускаем пустые строки
            data.append(row)

    return data


def import_from_csv(file, skip_first_row):
    """Импорт данных из CSV-файла"""
    data = []
    decoded_file = file.read().decode('utf-8').splitlines()
    reader = csv.reader(decoded_file)

    if skip_first_row:
        next(reader, None)

    for row in reader:
        if any(cell.strip() for cell in row):  # Пропускаем пустые строки
            data.append(row)

    return data


def export_medical_norms(format_type='xlsx', include_headers=True):
    """
    Экспорт норм медосмотров в файл

    Args:
        format_type: Формат экспорта ('xlsx', 'csv' или 'json')
        include_headers: Включать ли заголовки

    Returns:
        HttpResponse: Ответ для скачивания файла
    """
    norms = MedicalExaminationNorm.objects.select_related('harmful_factor', 'harmful_factor__examination_type').all()

    if format_type == 'xlsx':
        return export_to_xlsx(norms, include_headers)
    elif format_type == 'csv':
        return export_to_csv(norms, include_headers)
    elif format_type == 'json':
        return export_to_json(norms)
    else:
        # По умолчанию используем Excel
        return export_to_xlsx(norms, include_headers)


def export_to_xlsx(norms, include_headers):
    """Экспорт данных в Excel-файл"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Нормы медосмотров"

    # Добавляем заголовки
    if include_headers:
        headers = [
            'Наименование должности',
            'Код вредного фактора',
            'Наименование вредного фактора',
            'Вид медосмотра',
            'Периодичность (мес.)',
            'Переопределение периодичности',
            'Примечания'
        ]
        ws.append(headers)

    # Добавляем данные
    for norm in norms:
        row = [
            norm.position_name,
            norm.harmful_factor.short_name,
            norm.harmful_factor.full_name,
            norm.harmful_factor.examination_type.name,
            norm.harmful_factor.periodicity,
            norm.periodicity_override or '',
            norm.notes
        ]
        ws.append(row)

    # Автоматически регулируем ширину столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Создаем ответ
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"medical_norms_{timestamp}.xlsx"

    response = HttpResponse(buffer.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def export_to_csv(norms, include_headers):
    """Экспорт данных в CSV-файл"""
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"medical_norms_{timestamp}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Добавляем заголовки
    if include_headers:
        headers = [
            'Наименование должности',
            'Код вредного фактора',
            'Наименование вредного фактора',
            'Вид медосмотра',
            'Периодичность (мес.)',
            'Переопределение периодичности',
            'Примечания'
        ]
        writer.writerow(headers)

    # Добавляем данные
    for norm in norms:
        row = [
            norm.position_name,
            norm.harmful_factor.short_name,
            norm.harmful_factor.full_name,
            norm.harmful_factor.examination_type.name,
            norm.harmful_factor.periodicity,
            norm.periodicity_override or '',
            norm.notes
        ]
        writer.writerow(row)

    return response


def export_to_json(norms):
    """Экспорт данных в JSON-файл"""
    data = []

    for norm in norms:
        data.append({
            'position_name': norm.position_name,
            'harmful_factor': {
                'short_name': norm.harmful_factor.short_name,
                'full_name': norm.harmful_factor.full_name,
                'periodicity': norm.harmful_factor.periodicity,
            },
            'examination_type': norm.harmful_factor.examination_type.name,
            'periodicity_override': norm.periodicity_override,
            'notes': norm.notes
        })

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"medical_norms_{timestamp}.json"

    json_data = json.dumps(data, ensure_ascii=False, indent=4)

    response = HttpResponse(json_data, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def find_medical_norms_for_position(position, only_active=True):
    """
    Находит применимые нормы медосмотра для указанной должности

    Args:
        position: объект Position
        only_active: учитывать только активные нормы

    Returns:
        list: Список норм медосмотров для данной должности
    """
    # Ищем переопределения для конкретной должности
    position_factors = PositionMedicalFactor.objects.filter(position=position)

    if only_active:
        position_factors = position_factors.filter(is_disabled=False)

    # Сначала соберем все переопределения
    overrides_by_factor = {po.harmful_factor_id: po for po in position_factors}

    # Ищем эталонные нормы
    standard_norms = MedicalExaminationNorm.objects.filter(
        position_name__iexact=position.position_name
    ).select_related('harmful_factor', 'harmful_factor__examination_type')

    # Результирующий список норм
    result_norms = []

    # Добавляем переопределения
    for po in position_factors:
        if not only_active or not po.is_disabled:
            result_norms.append({
                'factor': po.harmful_factor,
                'examination_type': po.harmful_factor.examination_type,
                'periodicity': po.periodicity,
                'is_override': True,
                'is_disabled': po.is_disabled,
                'notes': po.notes
            })

    # Добавляем эталонные нормы, если для них нет переопределений
    for norm in standard_norms:
        if norm.harmful_factor_id not in overrides_by_factor:
            result_norms.append({
                'factor': norm.harmful_factor,
                'examination_type': norm.harmful_factor.examination_type,
                'periodicity': norm.periodicity,
                'is_override': False,
                'is_disabled': False,
                'notes': norm.notes
            })

    return result_norms


def get_employee_medical_examination_status(employee):
    """
    Получает статус медосмотров для сотрудника

    Args:
        employee: объект Employee

    Returns:
        dict: Информация о статусе медосмотров сотрудника
    """
    today = datetime.date.today()
    settings = MedicalSettings.get_settings(employee.organization)

    # Получаем все применимые виды медосмотров для должности сотрудника
    position = employee.position
    if not position:
        return {
            'status': 'unknown',
            'message': 'Не указана должность сотрудника',
            'medical_examinations': []
        }

    # Получаем все нормы для должности
    norms = find_medical_norms_for_position(position)
    if not norms:
        return {
            'status': 'not_required',
            'message': 'Медосмотр не требуется для данной должности',
            'medical_examinations': []
        }

    # Получаем последние медосмотры сотрудника по типам
    exams = EmployeeMedicalExamination.objects.filter(
        employee=employee
    ).order_by('examination_type', '-date_completed')

    last_exams_by_type = {}
    for exam in exams:
        if exam.examination_type_id not in last_exams_by_type:
            last_exams_by_type[exam.examination_type_id] = exam

    expired = []
    upcoming = []
    valid = []

    # Проверяем статус каждого требуемого медосмотра
    required_exam_types = set()
    for norm in norms:
        if norm['is_disabled']:
            continue

        exam_type = norm['examination_type']
        required_exam_types.add(exam_type.id)

        if exam_type.id in last_exams_by_type:
            exam = last_exams_by_type[exam_type.id]

            # Проверяем, истек ли срок
            if exam.next_date < today:
                expired.append({
                    'exam_type': exam_type,
                    'factor': norm['factor'],
                    'last_exam': exam,
                    'days_overdue': (today - exam.next_date).days
                })
            else:
                # Если остается менее days_before_issue дней, добавляем в upcoming
                days_remaining = (exam.next_date - today).days
                if days_remaining <= settings.days_before_issue:
                    upcoming.append({
                        'exam_type': exam_type,
                        'factor': norm['factor'],
                        'last_exam': exam,
                        'days_remaining': days_remaining
                    })
                else:
                    valid.append({
                        'exam_type': exam_type,
                        'factor': norm['factor'],
                        'last_exam': exam,
                        'days_remaining': days_remaining
                    })
        else:
            # Медосмотр не проходился
            expired.append({
                'exam_type': exam_type,
                'factor': norm['factor'],
                'last_exam': None,
                'days_overdue': 0  # Считаем, что просрочен с сегодняшнего дня
            })

    # Определяем общий статус
    status = 'valid'
    message = 'Все необходимые медосмотры пройдены'

    if expired:
        status = 'expired'
        if len(expired) == 1:
            message = f'Просрочен медосмотр: {expired[0]["exam_type"].name}'
        else:
            message = f'Просрочено {len(expired)} видов медосмотров'
    elif upcoming:
        status = 'upcoming'
        if len(upcoming) == 1:
            message = f'Скоро истекает срок медосмотра: {upcoming[0]["exam_type"].name} (осталось {upcoming[0]["days_remaining"]} дн.)'
        else:
            message = f'Скоро истекает срок {len(upcoming)} видов медосмотров'

    return {
        'status': status,
        'message': message,
        'expired': expired,
        'upcoming': upcoming,
        'valid': valid,
        'required_exam_types': required_exam_types
    }


def send_medical_examination_notifications():
    """
    Отправляет уведомления о приближающихся и просроченных медосмотрах.
    Обрабатывает каждую организацию отдельно с учетом её настроек.

    Returns:
        dict: Информация о количестве отправленных уведомлений
    """
    from directory.models import Organization
    today = timezone.now().date()

    total_sent = 0

    # Обрабатываем каждую организацию отдельно
    for organization in Organization.objects.all():
        settings = MedicalSettings.get_settings(organization)
        if not settings:
            continue

        # Определяем дату для отправки уведомлений
        notification_date = today + datetime.timedelta(days=settings.days_before_email)

        # Находим все медосмотры, требующие уведомления для этой организации
        to_notify_exams = EmployeeMedicalExamination.objects.filter(
            employee__organization=organization,
            next_date__lte=notification_date,
            next_date__gt=today,
            status='completed'
        ).select_related('employee', 'harmful_factor')

        # Здесь должна быть логика отправки уведомлений
        # Например, по email или в другие системы
        total_sent += to_notify_exams.count()

    return {
        'notifications_sent': total_sent,
        'timestamp': timezone.now()
    }